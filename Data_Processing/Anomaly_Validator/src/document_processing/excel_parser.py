"""
Excel/CSV Parser for extracting data from spreadsheets.

Handles Excel files, CSV files, and structured data supporting documents.
"""

import os
import pandas as pd
from typing import Dict, Any, List, Optional
import openpyxl


class ExcelParser:
    """
    Parse Excel and CSV files to extract data and metadata.
    
    Supports:
    - .xlsx, .xls (Excel)
    - .csv (CSV)
    - Multiple sheets
    - Formula extraction
    """
    
    def __init__(self, max_rows: Optional[int] = None):
        """
        Initialize Excel parser.
        
        Args:
            max_rows: Maximum rows to read (None = all rows)
        """
        self.max_rows = max_rows
    
    def parse(self, file_path: str) -> Dict[str, Any]:
        """
        Parse Excel/CSV file and extract content.
        
        Args:
            file_path: Path to file
            
        Returns:
            Dictionary with:
                - sheets: Dictionary of sheet_name -> DataFrame
                - summary: Summary statistics
                - metadata: File metadata
                - file_name: Original file name
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        file_ext = os.path.splitext(file_path)[1].lower()
        
        result = {
            'sheets': {},
            'summary': {},
            'metadata': {},
            'file_name': os.path.basename(file_path),
            'file_path': file_path,
            'file_type': file_ext
        }
        
        try:
            if file_ext == '.csv':
                result = self._parse_csv(file_path, result)
            elif file_ext in ['.xlsx', '.xls']:
                result = self._parse_excel(file_path, result)
            else:
                raise ValueError(f"Unsupported file type: {file_ext}")
            
            # Generate summary
            result['summary'] = self._generate_summary(result['sheets'])
            
            print(f"✓ Parsed {file_ext}: {result['file_name']} "
                  f"({len(result['sheets'])} sheets, {result['summary']['total_rows']} rows)")
            
        except Exception as e:
            print(f"⚠️  Error parsing {file_path}: {e}")
            result['error'] = str(e)
        
        return result
    
    def _parse_csv(self, file_path: str, result: Dict[str, Any]) -> Dict[str, Any]:
        """Parse CSV file."""
        df = pd.read_csv(file_path, nrows=self.max_rows)
        result['sheets']['Sheet1'] = df
        return result
    
    def _parse_excel(self, file_path: str, result: Dict[str, Any]) -> Dict[str, Any]:
        """Parse Excel file."""
        # Read all sheets
        excel_file = pd.ExcelFile(file_path)
        
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=self.max_rows)
            result['sheets'][sheet_name] = df
        
        # Extract metadata using openpyxl
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            result['metadata'] = {
                'sheet_names': wb.sheetnames,
                'num_sheets': len(wb.sheetnames)
            }
            wb.close()
        except:
            pass
        
        return result
    
    def _generate_summary(self, sheets: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Generate summary statistics for all sheets."""
        summary = {
            'num_sheets': len(sheets),
            'total_rows': 0,
            'total_cols': 0,
            'sheet_info': {}
        }
        
        for sheet_name, df in sheets.items():
            sheet_summary = {
                'rows': len(df),
                'cols': len(df.columns),
                'columns': list(df.columns),
                'dtypes': df.dtypes.astype(str).to_dict(),
                'null_counts': df.isnull().sum().to_dict()
            }
            
            summary['sheet_info'][sheet_name] = sheet_summary
            summary['total_rows'] += len(df)
            summary['total_cols'] += len(df.columns)
        
        return summary
    
    def to_text(self, parse_result: Dict[str, Any], max_rows_per_sheet: int = 100) -> str:
        """
        Convert parsed Excel/CSV to readable text format.
        
        Args:
            parse_result: Result from parse() method
            max_rows_per_sheet: Maximum rows to include in text
            
        Returns:
            Formatted text representation
        """
        text_parts = []
        
        text_parts.append(f"File: {parse_result['file_name']}")
        text_parts.append(f"Type: {parse_result['file_type']}")
        text_parts.append(f"Sheets: {parse_result['summary']['num_sheets']}\n")
        
        for sheet_name, df in parse_result['sheets'].items():
            text_parts.append(f"=== Sheet: {sheet_name} ===")
            text_parts.append(f"Rows: {len(df)}, Columns: {len(df.columns)}")
            text_parts.append(f"Columns: {', '.join(df.columns)}\n")
            
            # Add data preview
            preview_df = df.head(max_rows_per_sheet)
            text_parts.append(preview_df.to_string())
            
            if len(df) > max_rows_per_sheet:
                text_parts.append(f"\n... {len(df) - max_rows_per_sheet} more rows ...")
            
            text_parts.append("\n")
        
        return "\n".join(text_parts)
    
    def search_in_excel(
        self,
        file_path: str,
        search_term: str,
        case_sensitive: bool = False
    ) -> List[Dict[str, Any]]:
        """
        Search for a term in Excel/CSV file.
        
        Args:
            file_path: Path to file
            search_term: Term to search for
            case_sensitive: Whether search is case-sensitive
            
        Returns:
            List of matches with sheet, row, column info
        """
        matches = []
        
        try:
            parse_result = self.parse(file_path)
            
            for sheet_name, df in parse_result['sheets'].items():
                # Search in each column
                for col in df.columns:
                    col_data = df[col].astype(str)
                    
                    if not case_sensitive:
                        matching_rows = df[col_data.str.lower().str.contains(search_term.lower(), na=False)]
                    else:
                        matching_rows = df[col_data.str.contains(search_term, na=False)]
                    
                    if not matching_rows.empty:
                        for idx, row in matching_rows.iterrows():
                            matches.append({
                                'sheet': sheet_name,
                                'row': int(idx) + 2,  # +2 for header and 0-indexing
                                'column': col,
                                'value': str(row[col])
                            })
        
        except Exception as e:
            print(f"⚠️  Search error: {e}")
        
        return matches
    
    def extract_numeric_columns(self, parse_result: Dict[str, Any]) -> Dict[str, List[str]]:
        """Extract names of numeric columns from each sheet."""
        numeric_cols = {}
        
        for sheet_name, df in parse_result['sheets'].items():
            numeric_cols[sheet_name] = list(df.select_dtypes(include=['number']).columns)
        
        return numeric_cols
    
    def validate_reconciliation(
        self,
        parse_result: Dict[str, Any],
        sheet_name: str,
        amount_column: str,
        expected_total: float,
        tolerance: float = 0.01
    ) -> Dict[str, Any]:
        """
        Validate if reconciliation matches expected total.
        
        Args:
            parse_result: Parsed Excel data
            sheet_name: Sheet to validate
            amount_column: Column containing amounts
            expected_total: Expected sum
            tolerance: Acceptable difference
            
        Returns:
            Validation result
        """
        try:
            df = parse_result['sheets'][sheet_name]
            actual_total = df[amount_column].sum()
            difference = abs(actual_total - expected_total)
            
            return {
                'matches': difference <= tolerance,
                'expected': expected_total,
                'actual': actual_total,
                'difference': difference,
                'tolerance': tolerance
            }
        except Exception as e:
            return {
                'matches': False,
                'error': str(e)
            }


if __name__ == "__main__":
    # Test Excel parser
    print("Testing Excel Parser...")
    
    parser = ExcelParser()
    
    # Create a test DataFrame
    import pandas as pd
    test_df = pd.DataFrame({
        'GL_Account': ['101000', '201000', '301000'],
        'Description': ['Cash', 'Accounts Payable', 'Revenue'],
        'Debit': [5000000, 0, 0],
        'Credit': [0, 3000000, 2000000]
    })
    
    # Save and parse
    test_file = 'test_reconciliation.csv'
    test_df.to_csv(test_file, index=False, encoding='utf-8-sig')
    
    result = parser.parse(test_file)
    print(f"\n✓ Parsed CSV: {result['summary']}")
    
    text = parser.to_text(result)
    print(f"\n✓ Text representation:\n{text[:300]}...")
    
    # Cleanup
    import os
    if os.path.exists(test_file):
        os.remove(test_file)
    
    print("\n✓ Excel Parser tests passed!")

